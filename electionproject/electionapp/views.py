import csv
from django.contrib import messages
from django.shortcuts import render, redirect
from .models import *
from .forms import *
from django.http import HttpResponse
from openpyxl import load_workbook
from datetime import datetime
from django.db.models import Count


#Main Homepage
def index(request):
    return render(request, 'index.html')

def about(request):
    return render(request, 'about.html')

def contact(request):
    return render(request, 'contact.html')

def service(request):
    return render(request, 'services.html')

#Admin
def AdminView(request):
    uid=request.session['uid']
    return render(request, 'AdminU/adminhome.html',{'uid':uid})

def Admindistrictview(request):
    data = DistrictModel.objects.all()
    return render(request,'AdminU/admindistrictview.html',{'a':data})

def Adminlocalbodyview(request,id):
    a = LocalbodyModel.objects.filter(District_id=id)
    return render(request,"AdminU/adminlocalview.html",{'a':a})

def Adminwardview(request,id,District_id ):
    a = WardModel.objects.filter(Localbody_id=id,District_id=District_id)
    return render(request,"AdminU/adminwardview.html",{'a':a})

def Adminpollingview(request,id,District_id,Localbody_id):
    a = PollingstationModel.objects.filter(Ward_id=id,District_id=District_id,Localbody_id=Localbody_id)
    print(a)
    request.session['did']=District_id
    did=request.session['did']
    print(did)
    request.session['lid']=Localbody_id
    lid=request.session['lid']
    print(lid)
    request.session['wid']=id
    wid=request.session['wid']
    print(wid)
    return render(request,"AdminU/adminpollingview.html",{'a':a})

def AdminIssueUsers(request):
    return render(request,"AdminU/AdminIssueUsers.html")


# Admin Voters Display
def AdmindistrictDisplay(request):
    data = DistrictModel.objects.all()
    return render(request,'AdminU/districtview.html',{'a':data})

def AdmindistrictAdd(request):
    if request.method == 'POST':
        District = request.POST['District']
        data = DistrictModel(District=District)
        data.save()
        return redirect(AdmindistrictDisplay)
    return render(request, 'AdminU/districtadd.html')

def AdmindistrictUpdate(request,id):
    data = DistrictModel.objects.get(id=id)
    if request.method=='POST':
        data.District=request.POST.get("District")
        data.save()
        return redirect(AdmindistrictDisplay)
    return render(request,'AdminU/districtupdate.html',{'a':data})

def AdmindistrictDelete(request, id):
    data = DistrictModel.objects.get(id=id)
    data.delete()
    return redirect(AdmindistrictDisplay)


#localbody
def AdminlocalbodyDisplay(request,id):
    a = LocalbodyModel.objects.filter(District_id=id)
    return render(request,"AdminU/localbodyview.html",{'a':a})

def AdminlocalbodyAdd(request):
    if request.method == 'POST':
        Localbody_name = request.POST['Localbody_name']
        data = LocalbodyModel(Localbody_name=Localbody_name)
        data.save()
        return redirect(AdminlocalbodyDisplay)
    return render(request, 'AdminU/localbodyadd.html')

def AdminlocalbodyUpdate(request,id):
    data = LocalbodyModel.objects.get(id=id)
    if request.method=='POST':
        data.Localbody_name=request.POST.get("Localbody_name")
        data.save()
        return redirect(AdminlocalbodyDisplay)
    return render(request,'AdminU/localbodyupdate.html',{'a':data})

def AdminlocalbodyDelete(request, id):
    data = LocalbodyModel.objects.get(id=id)
    data.delete()
    return redirect(AdminlocalbodyDisplay)

#ward
def AdminwardDisplay(request,id,District_id ):
    a = WardModel.objects.filter(Localbody_id=id,District_id=District_id)
    return render(request,"AdminU/wardview.html",{'a':a})

def AdminwardAdd(request):
    if request.method == 'POST':
        Ward_name = request.POST['Ward_name']
        data = WardModel(Ward_name=Ward_name)
        data.save()
        return redirect(AdminwardDisplay)
    return render(request, 'AdminU/wardadd.html')

def AdminwardUpdate(request,id):
    data = WardModel.objects.get(id=id)
    if request.method=='POST':
        data.Ward_name=request.POST.get("Ward_name")
        data.save()
        return redirect(AdminwardDisplay)
    return render(request,'AdminU/wardupdate.html',{'a':data})

def AdminwardDelete(request, id):
    data = WardModel.objects.get(id=id)
    data.delete()
    return redirect(AdminwardDisplay)

#polling station
def AdminpollingDisplay(request,id,District_id,Localbody_id):
    a = PollingstationModel.objects.filter(Ward_id=id,District_id=District_id,Localbody_id=Localbody_id)
    print(a)
    request.session['did']=District_id
    did=request.session['did']
    print(did)
    request.session['lid']=Localbody_id
    lid=request.session['lid']
    print(lid)
    request.session['wid']=id
    wid=request.session['wid']
    print(wid)
    return render(request,"AdminU/pollingstation.html",{'a':a})

def AdminpollingAdd(request):
    if request.method == 'POST':
        Ward_name = request.POST['Pollingstation_name']
        data = PollingstationModel(Ward_name=Ward_name)
        data.save()
        return redirect(AdminpollingDisplay)
    return render(request, 'AdminU/pollingstationadd.html')

def AdminpollingUpdate(request,id):
    data = PollingstationModel.objects.get(id=id)
    if request.method=='POST':
        data.Pollingstation_name=request.POST.get("Pollingstation_name")
        data.save()
        return redirect(AdminpollingDisplay)
    return render(request,'AdminU/pollingstationupdate.html',{'a':data})

def AdminpollingDelete(request, id):
    data = PollingstationModel.objects.get(id=id)
    data.delete()
    return redirect(AdminpollingDisplay)

def AdminVoterDisplay(request,id):
    a = ExcelModel.objects.filter(Pollingstation_name=id)
    return render(request, 'AdminU/voterdisplay.html', {'a':a})

def AdminIssueDisplayLM(request):
    data = ReportIssueModel.objects.filter(privilege='LocalMember')
    return render(request, 'AdminU/issuedisplay.html', {'data':data})

def AdminIssueDisplayCM(request):
    data = ReportIssueModel.objects.filter(privilege='CounsilMember')
    return render(request, 'AdminU/issuedisplay.html', {'data':data})

def AdminIssueDisplaySM(request):
    data = ReportIssueModel.objects.filter(privilege='SuperMember')
    return render(request, 'AdminU/issuedisplay.html', {'data':data})

def AdminIssueDisplayP(request):
    data = ReportIssueModel.objects.filter(privilege='President')
    return render(request, 'AdminU/issuedisplay.html', {'data':data})

def AdminSuccess(request):
    return render(request, 'AdminU/successpage.html')

def AdminIssueReply(request,id):
    data=ReportIssueModel.objects.get(id=id)
    if request.method=='POST':
        reply=request.POST['reply']
        a=IssueReplyModel(username=data.username,userid=data.userid,datereceived=data.date,issue=data.issue,reply=reply)
        a.date=datetime.today()
        a.save()
        return redirect(AdminSuccess)
    return render(request,'AdminU/issuereply.html',{'data':data})

def IssueReplyDisplay(request):
    uid=request.session['uid']
    data = IssueReplyModel.objects.filter(username=uid)
    return render(request, 'issuereplydisplay.html', {'data':data})

def IssueReplyDisplayLM(request):
    uid=request.session['uid']
    data = IssueReplyModel.objects.filter(username=uid)
    return render(request, 'LocalMember/issuereplydisplay.html', {'data':data})

def IssueReplyDisplayCM(request):
    uid=request.session['uid']
    data = IssueReplyModel.objects.filter(username=uid)
    return render(request, 'CounsilMember/issuereplydisplay.html', {'data':data})

def IssueReplyDisplaySM(request):
    uid=request.session['uid']
    data = IssueReplyModel.objects.filter(username=uid)
    return render(request, 'SuperMember/issuereplydisplay.html', {'data':data})

def IssueReplyDisplayP(request):
    uid=request.session['uid']
    data = IssueReplyModel.objects.filter(username=uid)
    return render(request, 'President/issuereplydisplay.html', {'data':data})

def SuggestionDisplay(request):
    data = SuggestionModel.objects.all()
    return render(request, 'AdminU/suggestiondisplay.html', {'data':data})
        

#issue reporting
def IssueReport(request):
    uid=request.session['uid']
    uii=request.session['uii']
    if request.method=='POST':
        issue=request.POST['issue']
        # date=request.POST['date']
        data=ReportIssueModel(issue=issue,username=uid,userid=uii)
        data.date=datetime.today()
        data.save()
        return HttpResponse('Success')
    else:
        return render(request,'reportissue.html', {'uid':uid})

#suggessions  
def SuggesstionReport(request):
    uid=request.session['uid']
    uii=request.session['uii']
    if request.method=='POST':
        form = SuggestionForm(request.POST)
        if form.is_valid():
            suggestion=form.cleaned_data['suggestion']
            data=SuggestionModel(suggestion=suggestion,username=uid,userid=uii)
            data.date=datetime.today()
            data.save()
            return HttpResponse('Success')
        else:
            return HttpResponse('Failed')
    else:
        return render(request,'addsuggestion.html', {'uid':uid})


# User Registration.
def UserRegister(request):
    if request.method == 'POST':
        uid = request.POST['uid']
        username = request.POST['username']
        privilege = request.POST['privilege']
        firstname = request.POST['firstname']
        lastname = request.POST['lastname']
        email = request.POST['email']
        phone = request.POST['phone']
        district = request.POST['district']
        localbody = request.POST['localbody']
        ward = request.POST['ward']
        address = request.POST['address']
        password = request.POST['password']
        confirmpassword = request.POST['confirmpassword']
        if password == confirmpassword:
            b = UserRegisterModel(uid=uid, username=username, privilege=privilege, firstname=firstname,
                                  lastname=lastname, email=email, phone=phone, district=district, localbody=localbody,
                                  ward=ward, address=address, password=password, confirmpassword=confirmpassword)
            b.save()
            return redirect(UserLogin)
        else:
            return HttpResponse("Incorrect Password!")
    return render(request, 'User/signup.html')

# User Login.
def UserLogin(request):
    if request.method == 'POST':
        form = UserLoginForm(request.POST)
        if form.is_valid():
            username = form['username'].value()
            password = form['password'].value()
            try:
                request.session['uid']=username
          
                user = UserRegisterModel.objects.get(username=username, password=password, privilege='LocalMember')
                request.session['prv']=user.privilege  
                if user is not None:
                    uid=request.session['uid']
                    print(user.id)
                    request.session['uii']=user.id

                    messages.info(request, f"You are now logged in as {username}.")
                    return redirect(LocalMemberHome)
            except:
                pass

            try:
                user = UserRegisterModel.objects.get(username=username, password=password, privilege='CounsilMember')
                if user is not None:
                    cid=user.id
                    uid=request.session['uid']
                    messages.info(request, f"You are now logged in as {username}.")
                    # return redirect(CounsilerHome)
                    return render(request, 'CounsilMember/home.html', {'uid':uid, 'cid':cid})
            except:
                pass

            try:
                user = UserRegisterModel.objects.get(username=username, password=password, privilege='SuperMember')
                if user is not None:
                    uid=request.session['uid']
                    messages.info(request, f"You are now logged in as {username}.")
                    # return redirect(SuperMemberHome)
                    return render(request, 'SuperMember/home.html', {'uid':uid})
            except:
                pass

            try:
                user = UserRegisterModel.objects.get(username=username, password=password, privilege='President')
                if user is not None:
                    uid=request.session['uid']
                    messages.info(request, f"You are now logged in as {username}.")
                    # return redirect(PresidentHome)
                    return render(request, 'President/home4.html', {'uid':uid})
            except:
                pass

            try:
                user = UserRegisterModel.objects.get(username=username, password=password)
                if user is not None:
                    uid=request.session['uid']
                    messages.info(request, f"You are now logged in as {username}.")
                    return redirect(AdminView)
            except:
                pass
        else:
            return HttpResponse('Login Failed')
    return render(request, 'User/login.html')


def UserDisplay(request):
    data=UserRegisterModel.objects.all()
    return render(request,'AdminU/userdisplay.html',{'data':data})


def Supermember(request,id):
    c = CommentModel.objects.get(id=id)
    uid=request.session['uid']
    if request.method == 'POST':
        form = SuperMemberForm(request.POST)
        if form.is_valid():
            form.save()
        return redirect(SMSuccess)
    return render(request,'SuperMember/supermemberdata.html',{'c':c,'uid':uid})


def VerifyUser(request,id):
    data=UserRegisterModel.objects.get(id=id)
    uid=data.uid
    usr=data.username
    if request.method == 'POST':
        data.uid=request.POST.get('uid')
        data.username=request.POST.get('username')
        data.verify_user=request.POST.get('verify_user')
        data.save()
        return redirect(VerifiedUsers)
    return render(request,'AdminU/verifyusers.html',{'uid':uid,'usr':usr,'data':data})


def VerifiedUsers(request):
    data=UserRegisterModel.objects.filter(verify_user='Verified')
    return render(request,'AdminU/verifieddisplay.html',{'data':data})


def UserUpdate(request,id):
    data=UserRegisterModel.objects.get(id=id)
    if request.method=='POST':
        data.firstname=request.POST.get("firstname")
        data.lastname=request.POST.get("lastname")
        data.privilege=request.POST.get("privilege")
        data.email=request.POST.get("email")
        data.phone=request.POST.get("phone")
        data.district=request.POST.get("district")
        data.localbody=request.POST.get("localbody")
        data.ward=request.POST.get("ward")
        data.address=request.POST.get("address")
        data.save()
        return redirect(UserDisplay)
    return render(request,'AdminU/userupdate.html',{'data':data})


def UserDelete(request, id):
    data = UserRegisterModel.objects.get(id=id)
    data.delete()
    return redirect(UserDisplay)


#Local Member
def LocalMemberHome(request):
    uid=request.session['uid']
    return render(request, 'LocalMember/home.html', {'uid':uid})

#issue reporting
def LocalIssueReport(request):
    uid=request.session['uid']
    uii=request.session['uii']
    prv=request.session['prv']
    if request.method=='POST':
        issue=request.POST['issue']
        # date=request.POST['date']
        data=ReportIssueModel(issue=issue,username=uid,userid=uii,privilege=prv)
        data.date=datetime.today()
        data.save()
        return redirect(LMSuccess)
    else:
        return render(request,'LocalMember/reportissue.html', {'uid':uid})

#suggessions  
def LocalSuggesstionReport(request):
    uid=request.session['uid']
    uii=request.session['uii']
    prv=request.session['prv']
    if request.method=='POST':
        form = SuggestionForm(request.POST)
        if form.is_valid():
            suggestion=form.cleaned_data['suggestion']
            data=SuggestionModel(suggestion=suggestion,username=uid,userid=uii,privilege=prv)
            data.date=datetime.today()
            data.save()
            return redirect(LMSuccess)
        else:
            return HttpResponse('Failed')
    else:
        return render(request,'LocalMember/addsuggestion.html', {'uid':uid})
    
def VoterDisplay(request,id):
    a = ExcelModel.objects.filter(Pollingstation_name=id)
    return render(request, 'LocalMember/voterdisplay.html', {'a':a})

# Add Extra Details
def VoterDetails(request, id):
    a = ExcelModel.objects.get(id=id)
    uid = request.session['uid']
    if request.method == 'POST':
        phone = request.POST['phone']
        gmail = request.POST['gmail']
        instagram = request.POST['instagram']
        facebook = request.POST['facebook']
        linkedin = request.POST['linkedin']
        probability = request.POST['probability']
        data= AddSocialModel.objects.create(phone=phone, gmail=gmail, instagram=instagram, facebook=facebook, linkedin=linkedin, probability=probability,ExcelModel_id=id,userid=uid)
        data.save()
        return redirect(LMSuccess)
        # return render(request, 'voterdisplay.html', {'a':a})
    return render(request, 'LocalMember/voterdetails.html', {'a':a, 'uid':uid})

def LMSuccess(request):
    return render(request, 'LocalMember/successpage.html')


#Counsil Member
def CounsilMemberHome(request):
    uid=request.session['uid']
    return render(request, 'CounsilMember/home.html', {'uid':uid})

def CMSuccess(request):
    return render(request, 'CounsilMember/successpage.html')

#issue reporting
def CounsilIssueReport(request):
    uid=request.session['uid']
    uii=request.session['uii']
    if request.method=='POST':
        issue=request.POST['issue']
        # date=request.POST['date']
        data=ReportIssueModel(issue=issue,username=uid,userid=uii)
        data.date=datetime.today()
        data.save()
        return redirect(CMSuccess)
    else:
        return render(request,'CounsilMember/reportissue.html', {'uid':uid})

#suggessions  
def CounsilSuggesstionReport(request):
    uid=request.session['uid']
    uii=request.session['uii']
    if request.method=='POST':
        form = SuggestionForm(request.POST)
        if form.is_valid():
            suggestion=form.cleaned_data['suggestion']
            data=SuggestionModel(suggestion=suggestion,username=uid,userid=uii)
            data.date=datetime.today()
            data.save()
            return redirect(CMSuccess)
        else:
            return HttpResponse('Failed')
    else:
        return render(request,'CounsilMember/addsuggestion.html', {'uid':uid})
    
# Display Counselor
def display(request,id):
    a = AddSocialModel.objects.filter(ExcelModel_id=id)
    data = ExcelModel.objects.get(id=id)
    uid=data.idno
    nid=data.name
    iid=data.id
    return render(request, 'CounsilMember/display.html', {'a':a, 'uid':uid, 'nid':nid, 'iid':iid})
    

# Super Member
def SuperMemberHome(request):
    uid=request.session['uid']
    return render(request, 'SuperMember/home.html', {'uid':uid})


def SMSuccess(request):
    return render(request, 'SuperMember/successpage.html')


def SuperMemberdistrict(request):
    data = DistrictModel.objects.all()
    return render(request,'SuperMember/district.html',{'a':data})

def SuperMemberlocalbody(request,id):
    a = LocalbodyModel.objects.filter(District_id=id)
    return render(request,"SuperMember/local.html",{'a':a})

def SuperMemberward(request,id,District_id ):
    a = WardModel.objects.filter(Localbody_id=id,District_id=District_id)
    return render(request,"SuperMember/ward.html",{'a':a})

def SuperMemberpolling(request,id,District_id,Localbody_id):
    a = PollingstationModel.objects.filter(Ward_id=id,District_id=District_id,Localbody_id=Localbody_id)
    print(a)
    request.session['did']=District_id
    did=request.session['did']
    print(did)
    request.session['lid']=Localbody_id
    lid=request.session['lid']
    print(lid)
    request.session['wid']=id
    wid=request.session['wid']
    print(wid)
    return render(request,"SuperMember/polling.html",{'a':a})

def SuperMemberVoters(request,id):
    a = ExcelModel.objects.filter(Pollingstation_name=id)
    return render(request, 'SuperMember/voters.html', {'a':a})

#issue reporting
def SuperIssueReport(request):
    uid=request.session['uid']
    uii=request.session['uii']
    if request.method=='POST':
        issue=request.POST['issue']
        # date=request.POST['date']
        data=ReportIssueModel(issue=issue,username=uid,userid=uii)
        data.date=datetime.today()
        data.save()
        return redirect(SMSuccess)
    else:
        return render(request,'SuperMember/reportissue.html', {'uid':uid})

#suggessions  
def SuperSuggesstionReport(request):
    uid=request.session['uid']
    uii=request.session['uii']
    if request.method=='POST':
        form = SuggestionForm(request.POST)
        if form.is_valid():
            suggestion=form.cleaned_data['suggestion']
            data=SuggestionModel(suggestion=suggestion,username=uid,userid=uii)
            data.date=datetime.today()
            data.save()
            return redirect(SMSuccess)
        else:
            return HttpResponse('Failed')
    else:
        return render(request,'SuperMember/addsuggestion.html', {'uid':uid})
    

def SMUsersDisplay(request):
    data=UserRegisterModel.objects.filter(verify_user='Verified')
    return render(request,'SuperMember/assigntaskprofile.html',{'data':data})
    

def SMAssignTask(request,id):
    a=UserRegisterModel.objects.get(id=id)
    usr=a.username
    uid=request.session['uid']
    uii=a.uid
    if request.method=='POST':
        form = AssignTaskForm(request.POST)
        if form.is_valid():
            task=form.cleaned_data['task']
            status=form.cleaned_data['status']
            data=AssignTaskModel(task=task,username=uid,userid=uii,assignedto=usr,status=status)
            data.date=datetime.today()
            data.save()
            return redirect(SMSuccess)
        else:
            return HttpResponse('Failed')
    else:
        return render(request,'SuperMember/assigntask.html', {'a':a, 'uid':uid})
    

# President
def PresidentHome(request):
    uid=request.session['uid']
    return render(request, 'President/home4.html', {'uid':uid})

def PMSuccess(request):
    return render(request, 'President/successpage.html')

def PMDelete(request):
    return render(request, 'President/deletepage.html')

def Presidentdistrictview(request):
    data = DistrictModel.objects.all()
    return render(request,'President/presidentdistrict.html',{'a':data})

def Presidentlocalbodyview(request,id):
    a = LocalbodyModel.objects.filter(District_id=id)
    return render(request,"President/presidentlocal.html",{'a':a})

def Presidentwardview(request,id,District_id ):
    a = WardModel.objects.filter(Localbody_id=id,District_id=District_id)
    return render(request,"President/presidentward.html",{'a':a})

def Presidentpollingview(request,id,District_id,Localbody_id):
    a = PollingstationModel.objects.filter(Ward_id=id,District_id=District_id,Localbody_id=Localbody_id)
    print(a)
    request.session['did']=District_id
    did=request.session['did']
    print(did)
    request.session['lid']=Localbody_id
    lid=request.session['lid']
    print(lid)
    request.session['wid']=id
    wid=request.session['wid']
    print(wid)
    return render(request,"President/presidentpoll.html",{'a':a})

def PresidentVoterDisplay(request,id):
    a = ExcelModel.objects.filter(Pollingstation_name=id)
    return render(request, 'President/voterdisplay.html', {'a':a})


def PMdistrictview(request):
    data = DistrictModel.objects.all()
    return render(request,'President/pdistrict.html',{'a':data})

def PMlocalbodyview(request,id):
    a = LocalbodyModel.objects.filter(District_id=id)
    return render(request,"President/plb.html",{'a':a})

def PMwardview(request,id,District_id ):
    a = WardModel.objects.filter(Localbody_id=id,District_id=District_id)
    return render(request,"President/pward.html",{'a':a})

def PMpollingview(request,id,District_id,Localbody_id):
    a = PollingstationModel.objects.filter(Ward_id=id,District_id=District_id,Localbody_id=Localbody_id)
    print(a)
    request.session['did']=District_id
    did=request.session['did']
    print(did)
    request.session['lid']=Localbody_id
    lid=request.session['lid']
    print(lid)
    request.session['wid']=id
    wid=request.session['wid']
    print(wid)
    return render(request,"President/ppoll.html",{'a':a})

def PMVoterDisplay(request,id):
    a = ExcelModel.objects.filter(Pollingstation_name=id)
    return render(request, 'President/voterdisplay.html', {'a':a})

#issue reporting
def PresidentIssueReport(request):
    uid=request.session['uid']
    uii=request.session['uii']
    if request.method=='POST':
        issue=request.POST['issue']
        # date=request.POST['date']
        data=ReportIssueModel(issue=issue,username=uid,userid=uii)
        data.date=datetime.today()
        data.save()
        return redirect(PMSuccess)
    else:
        return render(request,'President/reportissue.html', {'uid':uid})

#suggessions  
def PresidentSuggesstionReport(request):
    uid=request.session['uid']
    uii=request.session['uii']
    if request.method=='POST':
        form = SuggestionForm(request.POST)
        if form.is_valid():
            suggestion=form.cleaned_data['suggestion']
            data=SuggestionModel(suggestion=suggestion,username=uid,userid=uii)
            data.date=datetime.today()
            data.save()
            return redirect(PMSuccess)
        else:
            return HttpResponse('Failed')
    else:
        return render(request,'President/addsuggestion.html', {'uid':uid})
    

def PresidentUsersDisplay(request):
    data=UserRegisterModel.objects.filter(verify_user='Verified')
    return render(request,'President/assigntaskprofile.html',{'data':data})
    

def PresidentAssignTask(request,id):
    a=UserRegisterModel.objects.get(id=id)
    usr=a.username
    uid=request.session['uid']
    uii=a.uid
    if request.method=='POST':
        form = AssignTaskForm(request.POST)
        if form.is_valid():
            task=form.cleaned_data['task']
            status=form.cleaned_data['status']
            data=AssignTaskModel(task=task,username=uid,userid=uii,assignedto=usr,status=status)
            data.date=datetime.today()
            data.save()
            return redirect(PMSuccess)
        else:
            return HttpResponse('Failed')
    else:
        return render(request,'President/assigntask.html', {'a':a, 'uid':uid})
    
def PresidentDisplayTask(request):
    data=AssignTaskModel.objects.all()
    return render(request,'President/viewtasks.html',{'data':data})
    
def PresidentTaskDelete(request, id):
    data = AssignTaskModel.objects.get(id=id)
    data.delete()
    return redirect(PMDelete)
    

# Filtering Voter's Constituencies
def Districtview(request):
    a = DistrictModel.objects.all
    return render(request,"LocalMember/districtview.html",{'a':a})

def Localbodyview(request,id):
    a = LocalbodyModel.objects.filter(District_id=id)
    return render(request,"LocalMember/localbodyview.html",{'a':a})


def Wardview(request,id,District_id ):
    a = WardModel.objects.filter(Localbody_id=id,District_id=District_id)
    return render(request,"LocalMember/wardview.html",{'a':a})


def pollingstationview(request,id,District_id,Localbody_id):
    a = PollingstationModel.objects.filter(Ward_id=id,District_id=District_id,Localbody_id=Localbody_id)
    return render(request,"LocalMember/pollingstation.html",{'a':a})


def addexceldetails(request,id,District_id,Localbody_id,Ward_id):
    a = ExcelModel.objects.filter(Pollingstation_id=id,District_id=District_id,Localbody_id=Localbody_id,Ward_id=Ward_id)
    # request.session['uid']=a.id
    return render(request,'addexceldetails.html',{'a':a})

# Extracting and adding data as excel sheets
def excel_import(request,id):
    data = PollingstationModel.objects.get(id=id)
    pid=data.id
    print(pid)
    did=request.session['did']
    lid=request.session['lid']
    wid=request.session['wid']

    if request.method == 'POST':
        form = ExcelForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']
            wb = load_workbook(excel_file)
            ws = wb.active
            for row in ws.iter_rows(min_row=0):
                serialno =  row[0].value
                name = row[1].value
                guardiansname = row[2].value
                houseno = row[3].value
                housename = row[4].value
                genderage = row[5].value
                idno = row[6].value
                excel = ExcelModel(serialno=serialno,name=name,guardiansname=guardiansname,houseno=houseno,housename=housename,genderage=genderage,idno=idno,Pollingstation_name=pid,District_name=did,Localbody_name=lid,Ward_name=wid)
                excel.save()
    else:
        form = ExcelForm()
    return render(request,'DataExtraction/excel_import.html',{'form':form, 'a':data})


def excelview(request,id,District_id,Localbody_id,Ward_id):
    a = ExcelModel.objects.filter( Pollingstation_id=id,District_id=District_id,Localbody_id=Localbody_id,Ward_id=Ward_id)
    # request.session['uid']=a.id
    return render(request,'DataExtraction/excelview.html',{'a':a})


#Counsil Member
def CounsilerHome(request):
    uid=request.session['uid']
    return render(request, 'CounsilMember/home2.html', {'uid':uid})

def Counsilerdistrictview(request):
    data = DistrictModel.objects.all()
    return render(request,'CounsilMember/counsilerdistrict.html',{'a':data})

def CounsilerLocalbodyview(request,id):
    a = LocalbodyModel.objects.filter(District_id=id)
    return render(request,"CounsilMember/counsilerlocal.html",{'a':a})

def CounsilerWardview(request,id,District_id ):
    a = WardModel.objects.filter(Localbody_id=id,District_id=District_id)
    return render(request,"CounsilMember/counsilerward.html",{'a':a})

def Counsilerpollingstationview(request,id,District_id,Localbody_id):
    a = PollingstationModel.objects.filter(Ward_id=id,District_id=District_id,Localbody_id=Localbody_id)
    return render(request,"CounsilMember/counsilerpoll.html",{'a':a})

def CounsilerVoterDisplay(request,id):
    a = ExcelModel.objects.filter(Pollingstation_name=id)
    return render(request, 'CounsilMember/counsilermenu.html', {'a':a})

def counsilerview(request):
    a = ExcelModel.objects.all()
    return render(request,'CounsilMember/counsilerdata.html',{'a':a})


def Comment(request, id):
    data = ExcelModel.objects.get(id=id)
    uid=request.session['uid']
    print(uid)
    uiu=data.name
    uii=data.idno
    gnm=data.guardiansname
    hnm=data.housename
    ga=data.genderage
    ps=data.Pollingstation_name
    print(data.id)
    print(uid)
    if request.method=='POST':
        choice=request.POST['choice']
        comments=request.POST['comments']
        b=CommentModel(choice=choice,comments=comments,name=uiu,guardiansname=gnm,housename=hnm,genderage=ga,idno=uii,userassigned=uid,Pollingstation_id=ps)
        b.save()
        return redirect(CMSuccess)
    else:
        return render(request,'CounsilMember/cmcomment.html', {'uid':uid, 'uii':uii, 'uiu':uiu})


    # if request.method=='POST':
    #     a = CommentForm(request.POST)
    #     if a.is_valid():
    #         choice=a.cleaned_data['choice']
    #         comments=a.cleaned_data['comments']
    #         b=CommentModel(choice=choice,comments=comments,idno=uii,name=uiu,guardiansname=data.guardiansname,housename=data.housename,genderage=data.genderage,userassigned=uid)
    #         b.save()
    #         return redirect('viewvoterdetials')
    #     else:
    #         return HttpResponse(request,'Failed')
    # else:
    #     return render(request,'CounsilMember/cmcomment.html', {'uid':uid, 'uii':uii, 'uiu':uiu})


def viewvoterdetials(request):
    voter=CommentModel.objects.all()
    return render(request,'CounsilMember/viewvoterdetails.html',{'voter':voter})


def display(request,id):
    a = AddSocialModel.objects.filter(ExcelModel_id=id)
    data = ExcelModel.objects.get(id=id)
    uid=data.idno
    nid=data.name
    iid=data.id
    return render(request, 'CounsilMember/display.html', {'a':a, 'uid':uid, 'nid':nid, 'iid':iid})


def Admincommentview(request):
    a = CommentModel.objects.all()
    return render(request,'Admin/admincommentview.html',{'a':a})


def Adminexcelview(request):
    b = ExcelModel.objects.all()
    return render(request,'Admin/adminexcelview.html',{'b':b})



def PresidentDataDisplay(request,id):
    data = CommentModel.objects.filter(Pollingstation_id=id,verify_user='Verified')
    # request.session['uid']=a.id
    return render(request,'President/displaypresident.html',{'data':data})


def Supermember(request,id):
    c = CommentModel.objects.get(id=id)
    uid=request.session['uid']
    if request.method == 'POST':
        c.choice = request.POST.get("choice")
        c.comments = request.POST.get("comments")
        c.name = request.POST.get("name")
        c.idno = request.POST.get("idno")
        c.verify_user = request.POST.get("verify_user")
        c.save()
        return redirect(SMSuccess)

        form = SuperMemberForm(request.POST)
        if form.is_valid():
            form.save()
        return HttpResponse('Added')
    return render(request,'SuperMember/supermemberdata.html',{'c':c,'uid':uid})


def Supermemberdistrictview(request):
    data = DistrictModel.objects.all()
    return render(request,'SuperMember/supermemberdistrict.html',{'a':data})


def SupermemberLocalview(request,id):
    a = LocalbodyModel.objects.filter(District_id=id)
    return render(request,"SuperMember/supermemberlocal.html",{'a':a})


def SupermemberWardview(request,id,District_id ):
    a = WardModel.objects.filter(Localbody_id=id,District_id=District_id)
    return render(request,"SuperMember/supermemberward.html",{'a':a})


def Supermemberpollview(request,id,District_id,Localbody_id):
    a = PollingstationModel.objects.filter(Ward_id=id,District_id=District_id,Localbody_id=Localbody_id)
    return render(request,"SuperMember/supermemberpoll.html",{'a':a})


def SupermemberVoterDisplay(request,id):
    a = ExcelModel.objects.filter(Pollingstation_name=id)
    return render(request, 'SuperMember/supermembermenu.html', {'a':a})


def SuperView(request,id):
    a = CommentModel.objects.filter(Pollingstation_id=id,choice='NO')
    uid=request.session['uid']
    print(uid)
    return render(request,'SuperMember/commentview.html',{'a':a, 'uid':uid})


def PresidentComment(request, id):
    c = CommentModel.objects.get(id=id)
    asg=c.userassigned
    uid=request.session['uid']
    if request.method=='POST':
        comment=request.POST['comment']
        assigned=request.POST['assigned']
        name=request.POST['name']
        idno = request.POST['idno']
        status=request.POST['status']
        b=PresidentModel(comment=comment,assigned=assigned,name=name,idno=idno,assignedto=asg,status=status)
        b.date=datetime.today()
        b.save()
        return redirect(PMSuccess)
    return render(request,'President/presidentcomment.html',{'c':c,'uid':uid, 'asg':asg})


def Count(request):
    yes=AddSocialModel.objects.filter(probability='YES').count()
    no=AddSocialModel.objects.filter(probability='NO').count()
    maybe=AddSocialModel.objects.filter(probability='MAYBE').count()
    # count=data.count("YES")
    print(yes)
    print(no)
    print(maybe)
    return render(request,'President/count.html', {'yes':yes,'maybe':maybe,'no':no})


def CountPercentage(request):
    # Get counts and percentages
    counts = AddSocialModel.objects.values('probability').annotate(count=Count('probability'))
    total = sum(c['count'] for c in counts)
    percentages = {
        c['probability']: (c['count'] / total) * 100
        for c in counts
    }

    context = {
        'percentages': percentages,
    }
    
    return render(request, 'AdminU/countpercentage.html', context)


def PresidentTaskDisplay(request):
    uid=request.session['uid']
    data = AssignTaskModel.objects.filter(assignedto=uid)
    return render(request, 'LocalMember/taskdisplay.html', {'data':data})

def TaskUpdate(request,id):
    data=AssignTaskModel.objects.get(id=id)
    if request.method=='POST':
        data.status = request.POST.get("status")
        data.save()
        return redirect(PresidentTaskDisplay)
    return render(request, 'LocalMember/taskupdate.html', {'data': data})

def PresidentTaskDisplayCM(request):
    uid=request.session['uid']
    data = AssignTaskModel.objects.filter(assignedto=uid)
    return render(request, 'CounsilMember/taskdisplay.html', {'data':data})

def TaskUpdateCM(request,id):
    data=AssignTaskModel.objects.get(id=id)
    if request.method=='POST':
        data.status = request.POST.get("status")
        data.save()
        return redirect(PresidentTaskDisplay)
    return render(request, 'CounsilMember/taskupdate.html', {'data': data})

def PresidentVoterTaskDisplayCM(request):
    uid=request.session['uid']
    data = PresidentModel.objects.filter(assignedto=uid)
    return render(request, 'CounsilMember/presidenttask.html', {'data':data})

def VoterTaskUpdateCM(request,id):
    data=PresidentModel.objects.get(id=id)
    if request.method=='POST':
        data.status = request.POST.get("status")
        data.save()
        return redirect(PresidentVoterTaskDisplayCM)
    return render(request, 'CounsilMember/votertaskupdate.html', {'data': data})




